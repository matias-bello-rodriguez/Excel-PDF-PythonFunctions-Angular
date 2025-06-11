import openpyxl
from openpyxl.drawing.image import Image

def debug_excel_content():
    try:
        # Cargar el archivo Excel
        wb = openpyxl.load_workbook('test_simple_design.xlsx', data_only=True)
        ws = wb.active
        
        print("=== DEBUG EXCEL CONTENT ===")
        print(f"Nombre de la hoja: {ws.title}")
        print(f"Dimensiones: {ws.max_row}x{ws.max_column}")
        
        # Encontrar headers
        print("\n=== HEADERS EN FILA 5 ===")
        headers = {}
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=5, column=col).value
            headers[col] = cell_value
            print(f"Columna {col} ({openpyxl.utils.get_column_letter(col)}): {cell_value}")
        
        # Buscar columna DESIGN
        design_column = None
        for col, header in headers.items():
            if header and 'DESIGN' in str(header).upper():
                design_column = col
                break
        
        print(f"\n=== COLUMNA DESIGN ENCONTRADA: {design_column} ===")
        if design_column:
            print(f"Letra de columna: {openpyxl.utils.get_column_letter(design_column)}")
            
            # Verificar contenido de las celdas en la columna design
            print("\n=== CONTENIDO COLUMNA DESIGN ===")
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=design_column)
                print(f"Fila {row}: valor='{cell.value}', tipo={type(cell.value)}")
        
        # Verificar imágenes
        print(f"\n=== IMÁGENES EN LA HOJA ===")
        if hasattr(ws, '_images'):
            images = ws._images
            print(f"Número de imágenes encontradas: {len(images)}")
            
            for i, img in enumerate(images):
                print(f"\nImagen {i+1}:")
                print(f"  - Anchor: {img.anchor}")
                if hasattr(img.anchor, '_from'):
                    anchor_from = img.anchor._from
                    print(f"  - Desde: Col={anchor_from.col}, Row={anchor_from.row}")
                    print(f"  - Letra columna: {openpyxl.utils.get_column_letter(anchor_from.col + 1)}")
                    
                    # Verificar si está en la columna design
                    if design_column and (anchor_from.col + 1) == design_column:
                        print(f"  - ¡ESTÁ EN COLUMNA DESIGN!")
                    else:
                        print(f"  - No está en columna design (design_col={design_column})")
                else:
                    print(f"  - Tipo de anchor: {type(img.anchor)}")
        else:
            print("No se encontraron imágenes o el atributo _images no está disponible")
        
        # Intentar con otro método para encontrar imágenes
        print(f"\n=== MÉTODO ALTERNATIVO PARA IMÁGENES ===")
        try:
            for drawing in ws._drawings:
                print(f"Drawing encontrado: {drawing}")
                if hasattr(drawing, 'images'):
                    for img in drawing.images:
                        print(f"  - Imagen: {img}")
        except Exception as e:
            print(f"Error con método alternativo: {e}")
            
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    debug_excel_content()

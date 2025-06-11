import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import io

# Crear workbook
wb = openpyxl.Workbook()
ws = wb.active

# Headers en fila 5
headers = ["UBICACIÓN", "CODE", "Width (m)", "Height (m)", "DESIGN", "Quantity"]
for i, header in enumerate(headers, 1):
    ws.cell(row=5, column=i, value=header)

# Datos de productos - IMPORTANTE: Dejar la columna DESIGN vacía (no poner ningún valor)
products = [
    ["SALON", "WIN001", 1.5, 1.2, None, 1],       # None en lugar de ""
    ["COCINA", "WIN002", 0.8, 1.0, None, 1],      # None en lugar de ""
    ["DORMITORIO", "WIN003", 1.2, 1.5, None, 2]   # None en lugar de ""
]

# Agregar datos
for i, product in enumerate(products, 6):
    for j, value in enumerate(product, 1):
        if value is not None:  # Solo escribir si no es None
            ws.cell(row=i, column=j, value=value)

# Crear imagen simple y agregarla
img_pil = PILImage.new('RGB', (100, 80), color='blue')
buffer = io.BytesIO()
img_pil.save(buffer, format='PNG')
buffer.seek(0)

# Agregar imagen a la columna DESIGN (E6) - primera fila de datos
img = Image(buffer)
img.width = 80
img.height = 60
img.anchor = 'E6'  # Columna E, fila 6
ws.add_image(img)

# Segunda imagen para la fila 7
img_pil2 = PILImage.new('RGB', (100, 80), color='green')
buffer2 = io.BytesIO()
img_pil2.save(buffer2, format='PNG')
buffer2.seek(0)

img2 = Image(buffer2)
img2.width = 80
img2.height = 60
img2.anchor = 'E7'  # Columna E, fila 7
ws.add_image(img2)

print("Creando archivo Excel limpio...")
wb.save('test_clean_design.xlsx')
print("Archivo creado: test_clean_design.xlsx")

# Verificar que se creó correctamente
print("\nVerificando contenido:")
wb_check = openpyxl.load_workbook('test_clean_design.xlsx', data_only=True)
ws_check = wb_check.active
for row in range(5, 9):
    for col in range(1, 7):
        cell_value = ws_check.cell(row=row, column=col).value
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"{col_letter}{row}: {cell_value} ({type(cell_value)})")
    print()

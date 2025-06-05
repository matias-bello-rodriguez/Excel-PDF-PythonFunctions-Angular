import openpyxl
from openpyxl.styles import PatternFill, Font

print("Creando archivo Excel de prueba...")

# Crear un nuevo workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Productos con Imágenes"

# Configurar headers
headers = ['WINDOW CODE', 'Location', 'Width (m)', 'Height (m)', 'Surface (m²)', 'Quantity']
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

# Datos de productos de ejemplo
products = [
    ['WIN-001', 'A01', 1.2, 1.5, 1.8, 2],
    ['WIN-002', 'A02', 1.0, 1.2, 1.2, 1],
    ['WIN-003', 'A03', 1.5, 2.0, 3.0, 3],
    ['WIN-004', 'B01', 0.8, 1.0, 0.8, 1],
    ['WIN-005', 'B02', 1.3, 1.8, 2.34, 2]
]

# Insertar datos de productos
for row_num, product in enumerate(products, 2):
    for col_num, value in enumerate(product, 1):
        ws.cell(row=row_num, column=col_num, value=value)

# Guardar el archivo
filename = "test_excel_simple.xlsx"
wb.save(filename)
print(f"Archivo Excel básico creado: {filename}")
print(f"Productos: {len(products)}")

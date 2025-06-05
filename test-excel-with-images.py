#!/usr/bin/env python3
"""
Script para crear un archivo Excel de prueba con imágenes incrustadas
para probar la funcionalidad de mapeo por posición.
"""

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font
import os
import requests
from io import BytesIO

def create_test_excel():
    # Crear un nuevo workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos con Imágenes"
    
    # Configurar headers
    headers = ['WINDOW CODE', 'Location', 'Width (m)', 'Height (m)', 'Surface (m²)', 'Quantity']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Datos de productos de ejemplo
    products = [
        ['WIN-001', 'A01', 1.2, 1.5, 1.8, 2],
        ['WIN-002', 'A02', 1.0, 1.2, 1.2, 1],
        ['WIN-003', 'A03', 1.5, 2.0, 3.0, 3],
        ['WIN-004', 'B01', 0.8, 1.0, 0.8, 1],
        ['WIN-005', 'B02', 1.3, 1.8, 2.34, 2]
    ]
    
    # Insertar datos de productos
    for row_num, product in enumerate(products, 2):
        for col_num, value in enumerate(product, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Crear imágenes de ejemplo (pequeños rectángulos de colores)
    try:
        from PIL import Image as PILImage, ImageDraw
        
        # Crear imágenes de ejemplo para cada producto
        for i, (code, location, _, _, _, _) in enumerate(products):
            # Crear una imagen simple con el código del producto
            img = PILImage.new('RGB', (200, 150), color=(100 + i*30, 150 + i*20, 200 - i*25))
            draw = ImageDraw.Draw(img)
            draw.text((50, 70), code, fill=(255, 255, 255))
            
            # Guardar la imagen temporalmente
            img_path = f"temp_img_{i}.png"
            img.save(img_path)
            
            # Insertar la imagen en Excel
            excel_img = Image(img_path)
            excel_img.width = 100
            excel_img.height = 75
            
            # Posicionar la imagen al lado del producto (columna G)
            ws.add_image(excel_img, f"G{i+2}")
            
            # Limpiar archivo temporal
            if os.path.exists(img_path):
                os.remove(img_path)
                
    except ImportError:
        print("PIL no está disponible, creando Excel sin imágenes de ejemplo")
        # Agregar nota sobre imágenes
        ws.cell(row=1, column=7, value="Imágenes")
        for i in range(len(products)):
            ws.cell(row=i+2, column=7, value=f"[Imagen {i+1}]")
    
    # Ajustar ancho de columnas
    column_widths = [12, 10, 12, 12, 12, 10, 15]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
    
    # Guardar el archivo
    filename = "test_excel_with_images.xlsx"
    wb.save(filename)
    print(f"Archivo Excel de prueba creado: {filename}")
    print(f"Productos: {len(products)}")
    print("Columnas: Código, Ubicación, Ancho, Alto, Superficie, Cantidad, Imágenes")
    
    return filename

if __name__ == "__main__":
    try:
        filename = create_test_excel()
        print(f"\n✅ Archivo Excel de prueba creado exitosamente: {filename}")
        print("Puedes usar este archivo para probar la funcionalidad de mapeo por posición.")
    except Exception as e:
        print(f"❌ Error al crear el archivo Excel: {e}")

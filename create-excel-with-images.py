#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para crear un archivo Excel con imagenes en posiciones especificas
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image, ImageDraw, ImageFont
import os

def create_sample_images():
    """Crear imagenes de muestra simples"""
    images_dir = "sample_images"
    os.makedirs(images_dir, exist_ok=True)
    
    # Crear 5 imagenes de muestra
    image_files = []
    colors = ['red', 'blue', 'green', 'orange', 'purple']
    
    for i, color in enumerate(colors, 1):
        # Crear imagen simple con color y texto
        img = Image.new('RGB', (200, 150), color=color)
        draw = ImageDraw.Draw(img)
        
        # Agregar texto
        try:
            # Intentar usar una fuente por defecto
            font = ImageFont.load_default()
        except:
            font = None
        
        text = f"WIN-00{i}"
        draw.text((10, 10), text, fill='white', font=font)
        draw.text((10, 30), f"Imagen {i}", fill='white', font=font)
        
        # Guardar imagen
        filename = f"{images_dir}/window_{i:03d}.png"
        img.save(filename)
        image_files.append(filename)
        print(f"Imagen creada: {filename}")
    
    return image_files

def create_excel_with_images():
    """Crear archivo Excel con imagenes insertadas"""
    print("Creando archivo Excel con imagenes...")
    
    # Crear imagenes de muestra
    image_files = create_sample_images()
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos con Imagenes"
    
    # Configurar headers
    headers = ['WINDOW CODE', 'Location', 'Width (m)', 'Height (m)', 'Surface (m2)', 'Quantity', 'Image']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Datos de productos
    products = [
        ['WIN-001', 'A01', 1.2, 1.5, 1.8, 2],
        ['WIN-002', 'A02', 1.0, 1.2, 1.2, 1],
        ['WIN-003', 'A03', 1.5, 2.0, 3.0, 3],
        ['WIN-004', 'B01', 0.8, 1.0, 0.8, 1],
        ['WIN-005', 'B02', 1.3, 1.8, 2.34, 2]
    ]
    
    # Insertar datos de productos
    for row_num, product in enumerate(products, 2):
        for col_num, value in enumerate(product, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Ajustar altura de las filas para las imagenes
    for row_num in range(2, len(products) + 2):
        ws.row_dimensions[row_num].height = 100
    
    # Ajustar ancho de la columna de imagenes
    ws.column_dimensions['G'].width = 20
    
    # Insertar imagenes en las celdas correspondientes
    for i, (image_file, product) in enumerate(zip(image_files, products)):
        if os.path.exists(image_file):
            try:
                # Crear objeto imagen de openpyxl
                img = ExcelImage(image_file)
                
                # Redimensionar imagen para que quepa en la celda
                img.width = 120
                img.height = 90
                
                # Calcular posicion de la celda (columna G, fila correspondiente al producto)
                row_num = i + 2  # +2 porque empezamos en fila 2 (fila 1 son headers)
                cell_ref = f"G{row_num}"
                
                # Agregar imagen a la hoja
                ws.add_image(img, cell_ref)
                
                print(f"Imagen {image_file} insertada en celda {cell_ref} para producto {product[0]}")
                
            except Exception as e:
                print(f"Error insertando imagen {image_file}: {str(e)}")
    
    # Guardar archivo
    filename = "test_excel_with_positioned_images.xlsx"
    wb.save(filename)
    print(f"\nArchivo Excel creado exitosamente: {filename}")
    print(f"Productos: {len(products)}")
    print(f"Imagenes insertadas: {len(image_files)}")
    
    return filename

if __name__ == "__main__":
    try:
        excel_file = create_excel_with_images()
        print(f"\nExito! Archivo creado exitosamente!")
        print(f"Archivo: {excel_file}")
        print(f"Las imagenes estan posicionadas en la columna G, filas 2-6")
        print(f"Cada imagen corresponde al producto en la misma fila")
    except Exception as e:
        print(f"Error creando el archivo: {str(e)}")
        import traceback
        traceback.print_exc()

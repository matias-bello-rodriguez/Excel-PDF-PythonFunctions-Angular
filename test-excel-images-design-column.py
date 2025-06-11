#!/usr/bin/env python3
"""
Script para crear un archivo Excel de prueba con imágenes únicamente en la columna DESIGN
"""

import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
import base64
from io import BytesIO
from PIL import Image as PILImage

def create_test_image(color, filename):
    """Crea una imagen de prueba pequeña con el color especificado"""
    img = PILImage.new('RGB', (100, 80), color=color)
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    
    # Guardar también como archivo
    img.save(filename)
    
    return buffer

def create_excel_with_design_images():
    """Crea un archivo Excel con imágenes solo en la columna DESIGN"""
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Images Design"
    
    # Headers en la fila 5 (como en el código)
    headers = [
        "UBICACIÓN", "CODE", "Width (m)", "Height (m)", "DESIGN", 
        "Quantity", "Surface (m²)", "Material", "Comments"
    ]
    
    # Agregar headers en la fila 5
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Datos de productos de prueba
    products = [
        ["SALON", "WIN001", 1.5, 1.2, "IMAGE", 1, 1.8, "Aluminio", "Ventana principal"],
        ["COCINA", "WIN002", 0.8, 1.0, "IMAGE", 1, 0.8, "PVC", "Ventana cocina"],
        ["DORMITORIO", "WIN003", 1.2, 1.5, "IMAGE", 2, 3.6, "Madera", "Ventanas dormitorio"],
        ["BAÑO", "WIN004", 0.6, 0.8, "IMAGE", 1, 0.48, "Aluminio", "Ventana baño"],
        ["LIVING", "WIN005", 2.0, 1.8, "IMAGE", 1, 3.6, "Aluminio", "Ventana living"]
    ]
    
    # Agregar datos de productos
    for row_idx, product in enumerate(products, 6):  # Empezar en fila 6
        for col_idx, value in enumerate(product, 1):
            if col_idx != 5:  # No escribir en la columna DESIGN (columna 5)
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Crear imágenes de prueba y agregarlas SOLO a la columna DESIGN
    colors = ['red', 'green', 'blue', 'yellow', 'purple']
    
    # Asegurar que existe el directorio de imágenes de muestra
    os.makedirs('sample_images', exist_ok=True)
    
    for idx, (product, color) in enumerate(zip(products, colors)):
        row_num = idx + 6  # Fila donde está el producto
        
        # Crear imagen de prueba
        image_filename = f'sample_images/design_{product[1]}.png'
        img_buffer = create_test_image(color, image_filename)
        
        # Agregar imagen a Excel en la columna DESIGN (columna E = 5)
        img = Image(img_buffer)
        img.width = 80
        img.height = 60
        
        # Anclar la imagen específicamente a la celda E{row_num} (columna DESIGN)
        img.anchor = f'E{row_num}'
        ws.add_image(img)
        
        print(f"Imagen agregada: {image_filename} en celda E{row_num} (fila {row_num}, columna 5)")
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 12  # UBICACIÓN
    ws.column_dimensions['B'].width = 10  # CODE
    ws.column_dimensions['C'].width = 12  # Width
    ws.column_dimensions['D'].width = 12  # Height
    ws.column_dimensions['E'].width = 15  # DESIGN
    ws.column_dimensions['F'].width = 10  # Quantity
    ws.column_dimensions['G'].width = 12  # Surface
    ws.column_dimensions['H'].width = 12  # Material
    ws.column_dimensions['I'].width = 20  # Comments
    
    # Guardar archivo
    filename = 'test_excel_design_column.xlsx'
    wb.save(filename)
    print(f"\n✅ Archivo Excel creado: {filename}")
    print(f"📊 {len(products)} productos con imágenes en la columna DESIGN")
    print(f"🖼️ {len(colors)} imágenes ancladas únicamente en la columna E (DESIGN)")
    
    return filename

if __name__ == "__main__":
    print("🔧 Creando archivo Excel de prueba con imágenes en columna DESIGN...")
    filename = create_excel_with_design_images()
    print(f"\n🎯 Archivo listo para probar el mapeo de imágenes: {filename}")
    print("\n📝 Instrucciones:")
    print("1. Inicia la aplicación Angular")
    print("2. Ve a la sección de importar Excel")
    print("3. Sube este archivo de prueba")
    print("4. Usa los botones de debug para verificar el mapeo")
    print("5. Las imágenes deberían aparecer solo en la columna DESIGN")

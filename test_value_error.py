import openpyxl

def test_excel_files():
    files_to_test = [
        'test_simple_design.xlsx',
        'test_clean_design.xlsx'
    ]
    
    for filename in files_to_test:
        try:
            print(f"\n=== ANALIZANDO {filename} ===")
            
            # Cargar con data_only=True para obtener valores evaluados
            wb = openpyxl.load_workbook(filename, data_only=True)
            ws = wb.active
            
            print("Contenido de la columna E (DESIGN):")
            for row in range(5, min(ws.max_row + 1, 10)):
                cell = ws.cell(row=row, column=5)
                cell_str = str(cell.value)
                print(f"  E{row}: valor='{cell.value}' tipo={type(cell.value)} str='{cell_str}'")
                
                # Detectar errores de Excel
                if cell_str.startswith('#'):
                    print(f"    *** ERROR DE EXCEL DETECTADO: {cell_str} ***")
                elif cell.value is None:
                    print(f"    *** CELDA VACÍA ***")
            
            # Verificar imágenes
            print(f"\nImágenes en {filename}:")
            if hasattr(ws, '_images'):
                images = ws._images
                print(f"  Número de imágenes: {len(images)}")
                for i, img in enumerate(images):
                    print(f"  Imagen {i+1}: anchor={img.anchor}")
            else:
                print("  No se detectaron imágenes")
                
        except Exception as e:
            print(f"Error procesando {filename}: {e}")

if __name__ == "__main__":
    test_excel_files()

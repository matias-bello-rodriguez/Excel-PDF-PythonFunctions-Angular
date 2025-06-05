#!/usr/bin/env python3
"""
Servicio Python para extraer imágenes de archivos Excel (.xlsx)
"""

import os
import base64
import zipfile
import tempfile
import shutil
import traceback
from typing import List, Dict, Any
from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

app = FastAPI(title="Excel Image Extractor", version="1.0.0")

# Configurar CORS para permitir conexiones desde Angular
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:4200", "http://127.0.0.1:4200"],
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE"],
    allow_headers=["*"],
)

class ExcelImageExtractor:
    """Clase para extraer imágenes de archivos Excel usando tanto openpyxl como ZIP"""
    
    def __init__(self):
        self.temp_dirs = []
    
    def extract_images_from_excel(self, excel_path: str, output_dir: str) -> List[Dict[str, Any]]:
        """
        Extrae imágenes de un archivo Excel combinando openpyxl y método ZIP
        """
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Archivo Excel no encontrado: {excel_path}")
        
        os.makedirs(output_dir, exist_ok=True)
        images_info = []
        
        try:
            # Método 1: Usar openpyxl para obtener información de posición
            position_info = self._extract_position_info_with_openpyxl(excel_path)
            
            # Método 2: Extraer imágenes del ZIP
            zip_images = self._extract_images_from_zip(excel_path, output_dir)
            
            # Combinar información de posición con imágenes extraídas
            images_info = self._merge_position_and_images(position_info, zip_images)
            
        except Exception as e:
            print(f"Error durante la extracción: {str(e)}")
            print(traceback.format_exc())
            
            # Fallback: solo extraer del ZIP sin información de posición
            try:
                images_info = self._extract_images_from_zip(excel_path, output_dir)
            except Exception as fallback_error:
                print(f"Error en fallback: {str(fallback_error)}")
                raise
        
        return images_info
    
    def _extract_position_info_with_openpyxl(self, excel_path: str) -> List[Dict[str, Any]]:
        """Extrae información de posición de imágenes usando openpyxl"""
        position_info = []
        
        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=False)
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Buscar imágenes en la hoja
                if hasattr(worksheet, '_images'):
                    for idx, img in enumerate(worksheet._images):
                        try:
                            # Obtener información de anclaje
                            anchor = img.anchor
                            cell_ref = None
                            row = None
                            column = None
                            column_letter = None
                            
                            if hasattr(anchor, '_from') and anchor._from:
                                cell_ref = f"{get_column_letter(anchor._from.col + 1)}{anchor._from.row + 1}"
                                row = anchor._from.row
                                column = anchor._from.col
                                column_letter = get_column_letter(anchor._from.col + 1)
                            
                            position_info.append({
                                'sheet': sheet_name,
                                'cell_address': cell_ref or 'unknown',
                                'row': row,
                                'column': column,
                                'column_letter': column_letter,
                                'anchor_type': type(anchor).__name__,
                                'image_index': idx
                            })
                            
                        except Exception as e:
                            print(f"Error procesando imagen {idx} en hoja {sheet_name}: {str(e)}")
                            continue
            
            workbook.close()
            
        except Exception as e:
            print(f"Error con openpyxl: {str(e)}")
        
        return position_info
    
    def _extract_images_from_zip(self, excel_path: str, output_dir: str) -> List[Dict[str, Any]]:
        """Extrae imágenes del archivo Excel tratándolo como ZIP"""
        images_info = []
        
        try:
            with zipfile.ZipFile(excel_path, 'r') as zip_file:
                # Buscar archivos de imagen en el directorio xl/media/
                for file_info in zip_file.filelist:
                    if file_info.filename.startswith('xl/media/') and not file_info.is_dir():
                        # Extraer el archivo
                        extracted_data = zip_file.read(file_info.filename)
                        
                        # Generar nombre de archivo único
                        original_name = os.path.basename(file_info.filename)
                        name, ext = os.path.splitext(original_name)
                        
                        # Crear archivo en el directorio de salida
                        output_path = os.path.join(output_dir, original_name)
                        counter = 1
                        while os.path.exists(output_path):
                            output_path = os.path.join(output_dir, f"{name}_{counter}{ext}")
                            counter += 1
                        
                        with open(output_path, 'wb') as output_file:
                            output_file.write(extracted_data)
                        
                        # Agregar información de la imagen
                        images_info.append({
                            'filename': os.path.basename(output_path),
                            'path': output_path,
                            'size': len(extracted_data),
                            'extension': ext.lower()
                        })
        
        except Exception as e:
            print(f"Error extrayendo del ZIP: {str(e)}")
            raise
        
        return images_info
    
    def _merge_position_and_images(self, position_info: List[Dict], zip_images: List[Dict]) -> List[Dict]:
        """Combina información de posición con imágenes extraídas"""
        merged_images = []
        
        # Copiar todas las imágenes del ZIP
        for img in zip_images:
            merged_images.append(img.copy())
        
        # Intentar asignar información de posición
        for i, position in enumerate(position_info):
            if i < len(merged_images):
                merged_images[i].update({
                    'sheet': position.get('sheet'),
                    'cell_address': position.get('cell_address'),
                    'row': position.get('row'),
                    'column': position.get('column'),
                    'column_letter': position.get('column_letter'),
                    'anchor_type': position.get('anchor_type')
                })
        
        return merged_images

@app.get("/")
async def root():
    """Endpoint raíz para verificar que el servidor esté funcionando"""
    return {
        "message": "Servidor de extracción de imágenes Excel funcionando",
        "version": "1.0.0",
        "status": "ok"
    }

@app.get("/health")
async def health_check():
    """Endpoint de verificación de salud"""
    return {
        "status": "healthy",
        "service": "excel-image-extractor",
        "openpyxl_available": True
    }

@app.post("/extract-images")
async def extract_images(file: UploadFile = File(...)):
    """
    Extrae imágenes de un archivo Excel subido
    """
    if not file.filename.endswith(('.xlsx', '.xlsm')):
        raise HTTPException(
            status_code=400, 
            detail="Solo se admiten archivos Excel (.xlsx, .xlsm)"
        )
    
    temp_dir = None
    
    try:
        # Crear directorio temporal
        temp_dir = tempfile.mkdtemp(prefix="excel_images_")
        excel_path = os.path.join(temp_dir, file.filename)
        
        # Guardar archivo temporal
        with open(excel_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        print(f"Archivo guardado temporalmente: {excel_path}")
        print(f"Tamaño del archivo: {len(content)} bytes")
        
        # Extraer imágenes
        output_dir = os.path.join(temp_dir, "images")
        extractor = ExcelImageExtractor()
        images_info = extractor.extract_images_from_excel(excel_path, output_dir)
        
        if not images_info:
            return {
                "success": True,
                "message": "No se encontraron imágenes en el archivo Excel",
                "images": [],
                "count": 0
            }
        
        # Preparar respuesta con imágenes en base64
        image_files = []
        for img_info in images_info:
            try:
                with open(img_info["path"], "rb") as img_file:
                    image_data = base64.b64encode(img_file.read()).decode("utf-8")
                    
                    # Determinar MIME type
                    mime_types = {
                        '.png': 'image/png',
                        '.jpg': 'image/jpeg',
                        '.jpeg': 'image/jpeg',
                        '.gif': 'image/gif',
                        '.bmp': 'image/bmp',
                        '.tiff': 'image/tiff',
                        '.svg': 'image/svg+xml'
                    }
                    
                    mime_type = mime_types.get(img_info["extension"], 'image/png')
                    
                    image_files.append({
                        "filename": img_info["filename"],
                        "data": image_data,
                        "mimeType": mime_type,
                        "size": img_info["size"],
                        "extension": img_info["extension"],
                        "sheet": img_info.get("sheet", "unknown"),
                        "cellAddress": img_info.get("cell_address", "unknown"),
                        "row": img_info.get("row"),
                        "column": img_info.get("column"),
                        "columnLetter": img_info.get("column_letter"),
                        "anchorType": img_info.get("anchor_type", "unknown")
                    })
                    
            except Exception as e:
                print(f"Error leyendo imagen {img_info['filename']}: {str(e)}")
                continue
        
        return {
            "success": True,
            "message": f"Se extrajeron {len(image_files)} imágenes correctamente",
            "images": image_files,
            "count": len(image_files)
        }
        
    except Exception as e:
        error_msg = f"Error procesando archivo: {str(e)}"
        print(error_msg)
        print(traceback.format_exc())
        
        raise HTTPException(status_code=500, detail=error_msg)
        
    finally:
        # Limpiar archivos temporales
        if temp_dir and os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
                print(f"Directorio temporal limpiado: {temp_dir}")
            except Exception as e:
                print(f"Error limpiando directorio temporal: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)

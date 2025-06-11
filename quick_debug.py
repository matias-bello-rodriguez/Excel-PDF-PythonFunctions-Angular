import openpyxl

wb = openpyxl.load_workbook('test_simple_design.xlsx', data_only=True)
ws = wb.active
print('=== ANÁLISIS DEL ARCHIVO ===')
print(f'Dimensiones: {ws.max_row}x{ws.max_column}')
print(f'Columna E (DESIGN) contenido:')
for row in range(1, min(ws.max_row + 1, 10)):
    cell = ws.cell(row=row, column=5)
    print(f'  Fila {row}: valor="{cell.value}", tipo={type(cell.value)}')
    if '#' in str(cell.value or ''):
        print(f'    *** CONTIENE #VALUE ERROR ***')

# Verificar si hay fórmulas en las celdas
print('\n=== VERIFICANDO FÓRMULAS ===')
wb_formulas = openpyxl.load_workbook('test_simple_design.xlsx', data_only=False)
ws_formulas = wb_formulas.active
for row in range(1, min(ws_formulas.max_row + 1, 10)):
    cell = ws_formulas.cell(row=row, column=5)
    if cell.value and str(cell.value).startswith('='):
        print(f'  Fila {row}: fórmula="{cell.value}"')
    elif cell.value:
        print(f'  Fila {row}: texto="{cell.value}"')

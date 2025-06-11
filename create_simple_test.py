import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import io

# Crear workbook
wb = openpyxl.Workbook()
ws = wb.active

# Headers en fila 5
headers = ["UBICACIÓN", "CODE", "Width (m)", "Height (m)", "DESIGN", "Quantity"]
for i, header in enumerate(headers, 1):
    ws.cell(row=5, column=i, value=header)

# Datos de productos
products = [
    ["SALON", "WIN001", 1.5, 1.2, "", 1],
    ["COCINA", "WIN002", 0.8, 1.0, "", 1],
    ["DORMITORIO", "WIN003", 1.2, 1.5, "", 2]
]

# Agregar datos
for i, product in enumerate(products, 6):
    for j, value in enumerate(product, 1):
        if j != 5:  # Skip DESIGN column
            ws.cell(row=i, column=j, value=value)

# Crear imagen simple y agregarla
img_pil = PILImage.new('RGB', (100, 80), color='red')
buffer = io.BytesIO()
img_pil.save(buffer, format='PNG')
buffer.seek(0)

# Agregar imagen a la columna DESIGN (E6)
img = Image(buffer)
img.width = 80
img.height = 60
img.anchor = 'E6'
ws.add_image(img)

print("Creando archivo Excel...")
wb.save('test_simple_design.xlsx')
print("Archivo creado: test_simple_design.xlsx")
